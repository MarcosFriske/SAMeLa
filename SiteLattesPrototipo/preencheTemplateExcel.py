import openpyxl
from openpyxl.styles import Alignment
from copy import copy
from typing import List, Dict, Any, Union
from pathlib import Path
import win32com.client as win32
import pythoncom
from PIL import Image
import tempfile
import os
from io import BytesIO
from fpdf import FPDF


class ExcelTemplatePreencher:
    """
    Classe responsável por preencher o template Excel com critérios, 
    substituir placeholders do cabeçalho, gerar imagem PNG e fragmentá-la em A4,
    e gerar PDF em memória pronto para envio via Flask.
    """

    def __init__(self, template_path: Union[str, Path]):
        self.template_path = Path(template_path)
        if not self.template_path.exists():
            raise FileNotFoundError(f"Template não encontrado em: {self.template_path}")
        self.wb = openpyxl.load_workbook(self.template_path)
        self.ws = self.wb.active

    # ===================================================
    # Substituição de placeholders no cabeçalho
    # ===================================================
    def substituir_placeholders(self, dados: Dict[str, str]) -> None:
        for row in self.ws.iter_rows(min_row=3, max_row=7, min_col=1, max_col=6):
            for cell in row:
                if isinstance(cell.value, str):
                    for placeholder, novo_valor in dados.items():
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, novo_valor)

    # ===================================================
    # Encontrar linha inicial dos critérios
    # ===================================================
    def encontrar_linha_inicial(self) -> int:
        for row in self.ws.iter_rows(min_row=9, max_row=self.ws.max_row):
            if isinstance(row[0].value, str) and "{{LINHA_INICIAL_CRITERIOS_ITEM}}" in row[0].value:
                return row[0].row
        raise ValueError("Linha inicial dos critérios não encontrada no template.")

    # ===================================================
    # Copiar formatação de linha
    # ===================================================
    def copiar_formatacao(self, origem: int, destino: int) -> None:
        for col in range(1, self.ws.max_column + 1):
            celula_origem = self.ws.cell(row=origem, column=col)
            celula_destino = self.ws.cell(row=destino, column=col)
            if celula_origem.coordinate not in self.ws.merged_cells:
                celula_destino.value = celula_origem.value
            celula_destino.font = copy(celula_origem.font)
            celula_destino.fill = copy(celula_origem.fill)
            celula_destino.border = copy(celula_origem.border)
            celula_destino.alignment = copy(celula_origem.alignment)
            celula_destino.number_format = celula_origem.number_format

    # ===================================================
    # Preencher critérios na planilha
    # ===================================================
    def preencher_criterios(self, criterios: List[Dict[str, Any]]) -> None:
        linha_inicial = self.encontrar_linha_inicial()
        qtd_novas = len(criterios)
        linha_primeira_movida = linha_inicial + 1

        # Desfazer merges abaixo da linha inicial
        for merged in list(self.ws.merged_cells.ranges):
            if merged.min_row >= linha_primeira_movida:
                self.ws.unmerge_cells(str(merged))

        # Inserir linhas adicionais
        if qtd_novas > 2:
            self.ws.insert_rows(idx=linha_inicial + 1, amount=qtd_novas - 2)

        # Preencher cada critério
        for i, criterio in enumerate(criterios):
            linha_atual = linha_inicial + i
            self.copiar_formatacao(linha_inicial, linha_atual)
            self.ws.cell(row=linha_atual, column=1, value=criterio["NUMERO"])
            self.ws.cell(row=linha_atual, column=2, value=criterio["CRITERIO"])
            self.ws.cell(row=linha_atual, column=3, value=criterio["PONTOS"])
            self.ws.cell(row=linha_atual, column=4, value=criterio["MAX_ITENS"])
            self.ws.cell(row=linha_atual, column=5, value=criterio["TOTAL_MAX"])
            formula = (
                f"=IF((E{linha_atual}*C{linha_atual})>D{linha_atual},"
                f"D{linha_atual},"
                f"(E{linha_atual}*C{linha_atual}))"
            )
            self.ws.cell(row=linha_atual, column=6, value=formula)

        # Somatório
        linha_somatorio = linha_inicial + qtd_novas
        linha_declaracao = linha_somatorio + 1
        self.ws.merge_cells(start_row=linha_somatorio, start_column=1, end_row=linha_somatorio, end_column=5)
        celula = self.ws.cell(row=linha_somatorio, column=1)
        celula.value = "Somatório dos pontos"
        celula.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.cell(row=linha_somatorio, column=6, value=f"=SUM(F{linha_inicial}:F{linha_somatorio - 1})")
        self.ws.merge_cells(start_row=linha_declaracao, start_column=1, end_row=linha_declaracao, end_column=6)
        celula_dec = self.ws.cell(row=linha_declaracao, column=1)
        celula_dec.alignment = Alignment(horizontal="center", vertical="center")

    # ===================================================
    # Salvar Excel em memória (BytesIO)
    # ===================================================
    def salvar_em_memoria(self) -> BytesIO:
        excel_bytes = BytesIO()
        self.wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes

    # ===================================================
    # Gerar imagem em memória usando Excel COM
    # ===================================================
    def gerar_imagem_em_memoria(self) -> Image.Image:
        pythoncom.CoInitialize()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            temp_path = tmp.name
        try:
            self.wb.save(temp_path)
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.Visible = False
            wb = excel.Workbooks.Open(str(temp_path))
            ws = wb.Sheets(1)

            ultima_linha = 1
            for row in range(1, ws.UsedRange.Rows.Count + 1):
                for col in range(1, 7):
                    if ws.Cells(row, col).Value not in (None, ""):
                        ultima_linha = max(ultima_linha, row)

            intervalo = f"A1:F{ultima_linha}"
            rng = ws.Range(intervalo)
            rng.CopyPicture(Format=win32.constants.xlBitmap)

            import PIL.ImageGrab as ImageGrab
            ws.Activate()
            excel.WindowState = win32.constants.xlMaximized
            import time; time.sleep(0.5)

            im = ImageGrab.grabclipboard()
            wb.Close(False)
            excel.Quit()
            if im is None:
                raise RuntimeError("Falha ao capturar imagem da planilha.")
            return im

        finally:
            os.remove(temp_path)
            pythoncom.CoUninitialize()

    # ===================================================
    # Fragmenta imagem em A4
    # ===================================================
    def gerar_fragmentos_a4(self, img: Image.Image, dpi=150, altura_logo_pt: int = 338) -> List[Image.Image]:
        """
        Fragmenta a imagem em páginas A4 considerando espaço da logo na primeira página.
        Ajuste mínimo para evitar perda de linhas e sem estourar memória.
        """
        largura_a4_mm = 210
        altura_a4_mm = 297
        largura_max = int(largura_a4_mm / 25.4 * dpi)
        altura_max = int(altura_a4_mm / 25.4 * dpi)
    
        # Redimensiona proporcionalmente à largura da página
        proporcao = largura_max / img.width
        nova_altura = int(img.height * proporcao)
        img_resized = img.resize((largura_max, nova_altura), Image.LANCZOS)
    
        fragmentos = []
        topo = 0
    
        # Primeira página: desconta altura da logo
        altura_util_primeira = altura_max - altura_logo_pt
        base = min(topo + altura_util_primeira, nova_altura)
        fragmentos.append(img_resized.crop((0, topo, largura_max, base)))
        topo = base
    
        # Demais páginas: usa A4 inteira
        while topo < nova_altura:
            base = min(topo + altura_max, nova_altura)
            fragmentos.append(img_resized.crop((0, topo, largura_max, base)))
            topo = base
    
        return fragmentos
    
    
    # ===================================================
    # Gerar PDF direto em memória a partir dos fragmentos
    # ===================================================
    @staticmethod
    def gerar_pdf_em_memoria(fragmentos: List[Image.Image], logo_path: Union[str, Path] = None,
                             altura_logo_pt: int = 100) -> BytesIO:
        """
        Gera PDF em memória a partir de fragmentos.
        Apenas primeira página considera logo.
        """
        pdf = FPDF(unit="pt", format="A4")
        largura_a4_pt, altura_a4_pt = 595, 842
    
        y_offset_primeira = altura_logo_pt if logo_path else 0
    
        for i, img in enumerate(fragmentos):
            pdf.add_page()
            y_offset = y_offset_primeira if i == 0 else 0
    
            if i == 0 and logo_path:
                logo_w = 100 * 1.6  # aumenta 60%
                logo_h = 0
                pdf.image(str(logo_path), x=(largura_a4_pt - logo_w) / 2, y=20, w=logo_w, h=logo_h)
    
            # Salvar fragmento temporário e inserir no PDF
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp_img:
                temp_path = tmp_img.name
                img.save(temp_path, format="PNG")
            try:
                pdf.image(temp_path, x=0, y=y_offset, w=largura_a4_pt)
            finally:
                os.remove(temp_path)
    
        pdf_bytes_str = pdf.output(dest='S').encode('latin1')
        pdf_bytes = BytesIO(pdf_bytes_str)
        pdf_bytes.seek(0)
        return pdf_bytes
